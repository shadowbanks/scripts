#!/usr/bin/env python3

import os
from time import sleep
import spotipy
from spotipy.oauth2 import SpotifyOAuth
from spotipy.exceptions import SpotifyException
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

CLIENT_ID = os.getenv("SPOTIPY_CLIENT_ID")
CLIENT_SECRET = os.getenv("SPOTIPY_CLIENT_SECRET")
REDIRECT_URI = os.getenv("SPOTIPY_REDIRECT_URI")

SCOPE = 'playlist-modify-public playlist-modify-private user-read-private'

user_id = "fztjehlovwqa7x8i3cy95ngnz"
playlist_endpoint="https://api.spotify.com/v1/users/" + user_id + "/playlists"

sp = spotipy.Spotify(auth_manager=SpotifyOAuth(
    client_id=CLIENT_ID,
    client_secret=CLIENT_SECRET,
    redirect_uri=REDIRECT_URI,
    scope=SCOPE
))

def main():
    user_profile = sp.current_user()
    if user_profile:
        user_id = user_profile['id']
    else:
        print("Failed to retrieve user profile.")
        return

    # Create a new playlist
    playlist_name = os.getenv("PLAYLIST_NAME", "Grind Season Only (GPT created list - still testing)")
    playlist_description = os.getenv("PLAYLIST_DESCRIPTION", "This playlist will be built to keep you locked in, confident, focused, and hungry. No heartbreaks, no soft moments — just straight 'I'm gonna make it' energy. PS: Spotify Playlist created by a script and GPT-4.0 provided the song list, still testing.")

    # Test playlist
    # playlist_name = "Test Playlist"
    # playlist_description = "This is a test playlist"
    playlist = sp.user_playlist_create(user=user_id,
                                       name=playlist_name,
                                       public=True,
                                       description=playlist_description)

    if not playlist:
        print("Failed to create playlist.")
        return
    playlist_id = playlist['id']
    print(f"Playlist '{playlist_name}' created successfully with ID: {playlist_id}")

    # Add tracks to the playlist

    # Read the CSV file
    path = "./data/spotify_playlist.xlsx"
    songs = readSongsFromExcel(path)

    if not songs:
        print("No songs found in the Excel file.")
        return
    
    # Get the URIs of the songs
    songs_uris = getSongUris(songs)

    if len(songs_uris) == 0:
        print("No valid song URIs found.")
        return
    
    # Add songs to the playlist
    addSongsToPlaylist(playlist_id, songs_uris)


def addSongsToPlaylist(playlist_id, songs_uris):
    if len(songs_uris) > 100:
        for i in range(0, len(songs_uris), 100):
            sp.playlist_add_items(playlist_id, songs_uris[i:i + 100])
            print(f"Added songs {i + 1} to {min(i + 100, len(songs_uris))} to the playlist.")
    else:
        sp.playlist_add_items(playlist_id, songs_uris)
        print(f"Added {len(songs_uris)} songs to the playlist.")


def getSongUris(songs):
    songs_uris = []
    missing_songs = []
    for song in songs:
        song_name, artist_name = song
        # Search for songs on Spotify
        try:
            results = sp.search(q=f"track:{song_name} artist:{artist_name}", type='track', limit=1)
            if not results or not results.get("tracks") or not results["tracks"]["items"]:
                missing_songs.append(song)
                continue
            items = results.get("tracks").get("items")
            if items:
                songs_uris.append(items[0].get("uri"))
            else:
                missing_songs.append(song)
        except SpotifyException as e:
            if e.http_status == 429:
                retry_after = int(e.headers.get('Retry-After', 5))
                print(f"Rate limit exceeded. Retrying after {retry_after} seconds...")
                sleep(retry_after)
                continue
        sleep(0.1)  # Sleep to avoid hitting rate limits too quickly
    if missing_songs:
        print("\n Songs not found on Spotify:")
        for t, a in missing_songs:
            print(f" - {t} by {a}")

    return songs_uris

def readSongsFromExcel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    songs = []
    if ws:
        # Get the list of songs from the Excel file
        i = 0
        for track_name, artist_name in ws.iter_rows( values_only=True):
            # print(track_name, artist_name)
            if isinstance(track_name, str) and isinstance(artist_name, str):
                songs.append((track_name.strip(), artist_name.strip()))
            else:
                print(f"Invalid data found in Excel: {track_name}, {artist_name}")
    songs.pop(0)  # Remove the header row if it exists
    return list(dict.fromkeys(songs))  # Remove duplicates while preserving order


# https://open.spotify.com/user/fztjehlovwqa7x8i3cy95ngnz

main()